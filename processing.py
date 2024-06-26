import openpyxl as xl

wb= xl.load_workbook("transaction.xlsx")

sheet= wb['Sheet1']

rows= sheet.max_row



for values in range(2, rows+1):
    wrong_prize= sheet.cell(values, 3)
    correct_prize= wrong_prize.value*0.9
    correct_prize_cell= sheet.cell(values,4)
    correct_prize_cell.value= correct_prize
    
wb.save('transaction2.xlsx')

    
